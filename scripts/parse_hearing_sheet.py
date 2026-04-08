#!/usr/bin/env python3
"""
記入済みヒアリングシートExcelを読み取り、テナントセットアップ情報をJSON出力する。

使い方:
  python scripts/parse_hearing_sheet.py <記入済みExcelファイルパス>

出力:
  JSON形式のテナント設定情報（provision_tenant.sh に渡すパラメータ + 車両/ユーザー登録SQL）

Claude Code連携:
  このスクリプトの出力をもとに、Claude Codeが provision_tenant.sh を実行し、
  さらに車両・ユーザーの登録SQLを生成・実行する。
"""

import sys
import json
import re
import openpyxl


def cell_val(ws, row, col):
    """セルの値を安全に取得。結合セルも考慮。"""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def is_placeholder(val):
    """プレースホルダー（ヒント文字列）かどうかを判定"""
    placeholders = [
        "該当する曜日に○", "例:", "おまかせ or",
        "紙 / Excel", "Wi-Fi /", "スマートフォン /",
        "iPhone /", "Chrome /", "あり / なし", "自由記述",
    ]
    for p in placeholders:
        if val.startswith(p):
            return True
    return False


def get_val(ws, row, col):
    """値を取得。プレースホルダーは空文字に。"""
    v = cell_val(ws, row, col)
    if is_placeholder(v):
        return ""
    return v


def find_section_rows(ws):
    """各セクションの開始行を特定"""
    sections = {}
    for row in range(1, ws.max_row + 1):
        val = cell_val(ws, row, 1)  # A列
        if val.startswith("1. 会社情報"):
            sections["company"] = row
        elif val.startswith("2. 車両情報"):
            sections["vehicles"] = row
        elif val.startswith("3. ユーザー情報"):
            sections["users"] = row
        elif val.startswith("4. システム設定"):
            sections["settings"] = row
        elif val.startswith("5. 現在の業務状況"):
            sections["operations"] = row
        elif val.startswith("6. 利用環境"):
            sections["environment"] = row
    return sections


def parse_hearing_sheet(filepath):
    """ヒアリングシートExcelを読み取り、構造化データを返す"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    sections = find_section_rows(ws)

    result = {
        "company": {},
        "vehicles": [],
        "users": [],
        "settings": {},
        "operations": {},
        "environment": {},
    }

    # === 会社情報 ===
    company_fields = {
        "会社名（正式名称）": "company_name",
        "代表者名": "representative",
        "郵便番号": "postal_code",
        "住所": "address",
        "電話番号": "phone",
        "FAX番号": "fax",
        "運行管理者氏名": "manager_name",
        "管理者メールアドレス": "manager_email",
        "事業許可番号": "license_number",
    }

    start = sections.get("company", 0)
    end = sections.get("vehicles", ws.max_row)
    for row in range(start, end):
        label = cell_val(ws, row, 2)  # B列
        if label in company_fields:
            result["company"][company_fields[label]] = get_val(ws, row, 4)  # D列

    # === 車両情報 ===
    start = sections.get("vehicles", 0)
    end = sections.get("users", ws.max_row)
    for row in range(start, end):
        a_val = ws.cell(row=row, column=1).value
        if isinstance(a_val, (int, float)) and 1 <= a_val <= 99:
            plate = get_val(ws, row, 2)   # B列
            model = get_val(ws, row, 4)   # D列
            note = get_val(ws, row, 5)    # E列
            if plate or model:
                result["vehicles"].append({
                    "plate_number": plate,
                    "model": model,
                    "note": note,
                })

    # === ユーザー情報 ===
    start = sections.get("users", 0)
    end = sections.get("settings", ws.max_row)
    for row in range(start, end):
        a_val = ws.cell(row=row, column=1).value
        if isinstance(a_val, (int, float)) and 1 <= a_val <= 99:
            name = get_val(ws, row, 2)      # B列
            login_id = get_val(ws, row, 4)  # D列
            if not name:
                continue

            circle_values = ["○", "〇", "O", "o", "◯", "●", "Yes", "yes", "1", "TRUE"]
            is_driver = get_val(ws, row, 5) in circle_values
            is_inspector = get_val(ws, row, 6) in circle_values
            is_admin = get_val(ws, row, 7) in circle_values

            result["users"].append({
                "name": name,
                "login_id": login_id,
                "is_driver": is_driver,
                "is_inspector": is_inspector,
                "is_admin": is_admin,
                "permission_level": "Admin" if is_admin else "User",
            })

    # === システム設定 ===
    settings_fields = {
        "営業日": "business_days",
        "営業時間": "business_hours",
        "テーマカラー": "theme_color",
    }
    start = sections.get("settings", 0)
    end = sections.get("operations", ws.max_row)
    for row in range(start, end):
        label = cell_val(ws, row, 2)
        if label in settings_fields:
            result["settings"][settings_fields[label]] = get_val(ws, row, 4)

    # === 現在の業務状況 ===
    operations_fields = {
        "現在の日常点検の方法は？": "inspection_method",
        "現在の運行記録の方法は？": "record_method",
        "現在の予約管理の方法は？": "reservation_method",
        "既存の帳票フォーマットがあれば添付してください": "has_existing_forms",
        "過去の監査で指摘された事項があれば記入": "audit_notes",
    }
    start = sections.get("operations", 0)
    end = sections.get("environment", ws.max_row)
    for row in range(start, end):
        label = cell_val(ws, row, 2)
        if label in operations_fields:
            result["operations"][operations_fields[label]] = get_val(ws, row, 5)

    # === 利用環境 ===
    env_fields = {
        "事務所のインターネット環境": "internet",
        "運転者が使用する端末": "device",
        "端末のOS": "device_os",
        "ブラウザ": "browser",
    }
    start = sections.get("environment", 0)
    for row in range(start, ws.max_row + 1):
        label = cell_val(ws, row, 2)
        if label in env_fields:
            val = get_val(ws, row, 5)
            if not val:
                val = get_val(ws, row, 4)
            result["environment"][env_fields[label]] = val

    return result


def _company_to_prefix(company_name):
    """会社名からログインIDプレフィックスを生成。
    例: '株式会社テスト介護タクシー' → 'test'
         '近畿福祉輸送' → 'kinki'
         'ABC Transport' → 'abc'
    """
    # 「株式会社」「有限会社」等を除去
    name = re.sub(r'(株式会社|有限会社|合同会社|一般社団法人|NPO法人)', '', company_name).strip()

    # 英数字のみ抽出（英語社名の場合）
    ascii_part = re.sub(r'[^a-zA-Z0-9]', '', name)
    if len(ascii_part) >= 3:
        return ascii_part[:10].lower()

    # 日本語名の場合はテナントIDとして使えるローマ字が必要
    # → プレフィックスを空にして、tenant_id をプレフィックスに使う
    return ""


def _auto_generate_login_ids(users, company_name, tenant_id=""):
    """ログインIDが空のユーザーに自動生成する。
    パターン: {prefix}{番号:02d} (例: smiley01, smiley02)
    """
    prefix = _company_to_prefix(company_name)
    if not prefix and tenant_id:
        prefix = tenant_id
    if not prefix:
        prefix = "user"

    # 既に指定されたIDを収集（重複防止）
    used_ids = {u["login_id"] for u in users if u["login_id"]}

    counter = 1
    for user in users:
        if not user["login_id"]:
            while True:
                generated = f"{prefix}{counter:02d}"
                if generated not in used_ids:
                    break
                counter += 1
            user["login_id"] = generated
            user["login_id_auto"] = True
            used_ids.add(generated)
            counter += 1

    return users


def generate_provision_params(data, tenant_id=""):
    """パース結果からプロビジョニングパラメータを生成"""
    company = data.get("company", {})
    users = data.get("users", [])

    # ログインIDが空のユーザーに自動生成
    company_name = company.get("company_name", "")
    users = _auto_generate_login_ids(users, company_name, tenant_id)

    params = {
        "tenant_id": "<<要指定: テナントID（英数字）>>",
        "db_name": "<<要指定: DB名>>",
        "base_path": "<<要指定: ベースパス（例: /wts-tenants/xxx）>>",
        "system_name": "スマルト",
        "company_info": company,
        "vehicles": data.get("vehicles", []),
        "users": users,
        "theme_color": data.get("settings", {}).get("theme_color", ""),
        "settings": data.get("settings", {}),
    }

    return params


def generate_company_json(data, output_path):
    """company_info テーブル投入用のJSONファイルを生成する。
    provision_tenant.sh の第5引数として渡せる形式。"""
    company = data.get("company", {})
    company_json = {
        "company_name": company.get("company_name", ""),
        "representative": company.get("representative", ""),
        "postal_code": company.get("postal_code", ""),
        "address": company.get("address", ""),
        "phone": company.get("phone", ""),
        "fax": company.get("fax", ""),
        "manager_name": company.get("manager_name", ""),
        "manager_email": company.get("manager_email", ""),
        "license_number": company.get("license_number", ""),
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(company_json, f, ensure_ascii=False, indent=2)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("使い方: python scripts/parse_hearing_sheet.py <Excelパス> [--tenant-id <ID>] [--export-company <出力先.json>]", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]

    # --tenant-id オプション（ログインID自動生成のプレフィックスに使用）
    tenant_id = ""
    if "--tenant-id" in sys.argv:
        idx = sys.argv.index("--tenant-id")
        if idx + 1 < len(sys.argv):
            tenant_id = sys.argv[idx + 1]

    # --export-company オプション
    export_company_path = None
    if "--export-company" in sys.argv:
        idx = sys.argv.index("--export-company")
        if idx + 1 < len(sys.argv):
            export_company_path = sys.argv[idx + 1]

    data = parse_hearing_sheet(filepath)
    params = generate_provision_params(data, tenant_id)

    output = {
        "parsed_data": data,
        "provision_params": params,
    }

    # UTF-8で出力
    sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps(output, ensure_ascii=False, indent=2))

    # company_info JSON出力
    if export_company_path:
        generate_company_json(data, export_company_path)
        print(f"\ncompany_info JSON出力: {export_company_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
