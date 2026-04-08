#!/usr/bin/env python3
"""
スマルト導入ヒアリングシート Excel生成スクリプト

生成先: docs/hearing-sheet.xlsx
クライアントに送付し、記入後にClaude Codeで読み取ってテナントセットアップに使用する。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# === スタイル定義 ===
BLUE = "0D6EFD"
LIGHT_BLUE = "E8F0FE"
WHITE = "FFFFFF"
GRAY = "F5F5F5"
RED = "DC3545"

font_title = Font(name="Yu Gothic", size=18, bold=True, color=BLUE)
font_subtitle = Font(name="Yu Gothic", size=10, color="888888")
font_section = Font(name="Yu Gothic", size=13, bold=True, color=BLUE)
font_header = Font(name="Yu Gothic", size=10, bold=True, color="1A1A1A")
font_normal = Font(name="Yu Gothic", size=10, color="1A1A1A")
font_note = Font(name="Yu Gothic", size=9, color="666666")
font_input = Font(name="Yu Gothic", size=10, color="333333")
font_required = Font(name="Yu Gothic", size=9, bold=True, color=WHITE)
font_optional = Font(name="Yu Gothic", size=9, bold=True, color=WHITE)

fill_header = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_input = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_required = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
fill_optional = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
fill_blue = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
fill_light_gray = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_hearing_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ヒアリングシート"

    # 列幅設定
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    row = 1

    # === タイトル ===
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="スマルト 導入ヒアリングシート")
    cell.font = font_title
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="Welfare Transport System - Setup Hearing Sheet")
    cell.font = font_subtitle
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # 記入日
    row += 1
    ws.cell(row=row, column=4, value="記入日:").font = font_normal
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=5, value="").font = font_input
    ws.cell(row=row, column=5).border = Border(bottom=Side(style="thin", color="999999"))
    ws.cell(row=row, column=5).number_format = "YYYY/MM/DD"
    ws.cell(row=row, column=5).alignment = align_center
    row += 2

    # ============================================================
    # セクション1: 会社情報
    # ============================================================
    row = _write_section_header(ws, row, "1. 会社情報", required=True)

    company_fields = [
        "会社名（正式名称）",
        "代表者名",
        "郵便番号",
        "住所",
        "電話番号",
        "FAX番号",
        "運行管理者氏名",
        "管理者メールアドレス",
        "事業許可番号",
    ]
    for field in company_fields:
        row = _write_form_row(ws, row, field)

    row = _write_note(ws, row, "※ 事業許可番号は一般乗用旅客自動車運送事業の許可番号をご記入ください。")
    row += 1

    # ============================================================
    # セクション2: 車両情報
    # ============================================================
    row = _write_section_header(ws, row, "2. 車両情報", required=True)

    # ヘッダー
    headers = ["No.", "車両番号（ナンバー）", "車種名", "備考"]
    cols = [1, 2, 4, 5]
    merge_ranges = [None, "B{r}:C{r}", None, "E{r}:G{r}"]

    for i, (header, col) in enumerate(zip(headers, cols)):
        if merge_ranges[i]:
            ws.merge_cells(merge_ranges[i].format(r=row))
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = thin_border
        cell.alignment = align_center

    # No.の右セルにもボーダー
    ws.cell(row=row, column=1).border = thin_border
    # 結合セルの右端にもボーダー
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    for i in range(1, 7):
        ws.cell(row=row, column=1, value=i).font = font_normal
        ws.cell(row=row, column=1).alignment = align_center
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row=row, column=2).font = font_input
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4).font = font_input
        ws.cell(row=row, column=4).border = thin_border
        ws.merge_cells(f"E{row}:G{row}")
        ws.cell(row=row, column=5).font = font_input
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=7).border = thin_border
        row += 1

    row += 1

    # ============================================================
    # セクション3: ユーザー情報
    # ============================================================
    row = _write_section_header(ws, row, "3. ユーザー情報", required=True)

    # ヘッダー
    user_headers = ["No.", "氏名", "ログインID（任意）", "運転者", "点呼者", "管理者"]
    user_cols = [1, 2, 4, 5, 6, 7]
    user_merges = [None, "B{r}:C{r}", None, None, None, None]

    for i, (header, col) in enumerate(zip(user_headers, user_cols)):
        if user_merges[i]:
            ws.merge_cells(user_merges[i].format(r=row))
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = thin_border
        cell.alignment = align_center
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    for i in range(1, 9):
        ws.cell(row=row, column=1, value=i).font = font_normal
        ws.cell(row=row, column=1).alignment = align_center
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row=row, column=2).font = font_input
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4).font = font_input
        ws.cell(row=row, column=4).border = thin_border
        # 役割チェック（○を入力してもらう）
        for c in [5, 6, 7]:
            ws.cell(row=row, column=c).font = font_input
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).alignment = align_center
        row += 1

    row = _write_note(ws, row, "※ 該当する役割の列に「○」を入力してください。1人が複数の役割を兼ねることができます。")
    row = _write_note(ws, row, "※ ログインIDは空欄の場合、会社名から自動生成します（例: smiley01, smiley02 ...）。")
    row += 1

    # 役割説明テーブル
    role_header_row = row
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="役割").font = font_header
    ws.cell(row=row, column=2).fill = fill_light_gray
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=3).border = thin_border
    ws.merge_cells(f"D{row}:G{row}")
    ws.cell(row=row, column=4, value="説明").font = font_header
    ws.cell(row=row, column=4).fill = fill_light_gray
    ws.cell(row=row, column=4).border = thin_border
    for c in [5, 6, 7]:
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    roles = [
        ("運転者", "車両を運転する方。日常点検・出庫/入庫・乗車記録・売上金確認を行います。"),
        ("点呼者", "乗務前・乗務後の点呼を実施する方。運行管理者や運行管理補助者が該当します。"),
        ("管理者", "システム全体の設定変更、ユーザー・車両の追加/編集、全データの閲覧ができます。"),
    ]
    for role_name, role_desc in roles:
        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row=row, column=2, value=role_name).font = Font(name="Yu Gothic", size=9, bold=True)
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.merge_cells(f"D{row}:G{row}")
        ws.cell(row=row, column=4, value=role_desc).font = Font(name="Yu Gothic", size=9)
        ws.cell(row=row, column=4).alignment = align_left
        ws.cell(row=row, column=4).border = thin_border
        for c in [5, 6, 7]:
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    row = _write_note(ws, row, "※ 例: 少人数の事業所では、1人が「運転者 + 点呼者 + 管理者」を兼ねるケースもあります。")
    row += 1

    # ============================================================
    # セクション4: システム設定
    # ============================================================
    row = _write_section_header(ws, row, "4. システム設定", required=False)

    row = _write_form_row(ws, row, "営業日", "該当する曜日に○: 月 火 水 木 金 土 日")
    row = _write_form_row(ws, row, "営業時間", "例: 8:00 〜 18:00")
    row = _write_form_row(ws, row, "テーマカラー", "おまかせ or カラーコード（例: #00C896）")

    row = _write_note(ws, row, "※ テーマカラーは「おまかせ」の場合、スマルト標準色を適用します。")
    row += 1

    # ============================================================
    # セクション5: 現在の業務状況
    # ============================================================
    row = _write_section_header(ws, row, "5. 現在の業務状況", required=False)

    questions = [
        ("現在の日常点検の方法は？", "紙 / Excel / その他"),
        ("現在の運行記録の方法は？", "紙 / Excel / その他"),
        ("現在の予約管理の方法は？", "紙 / Excel / ホワイトボード / その他"),
        ("既存の帳票フォーマットがあれば添付してください", "あり / なし"),
        ("過去の監査で指摘された事項があれば記入", "自由記述"),
    ]
    for q, hint in questions:
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=q).font = font_normal
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4).border = thin_border
        ws.merge_cells(f"E{row}:G{row}")
        ws.cell(row=row, column=5).font = font_input
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=7).border = thin_border
        # プレースホルダーコメント
        ws.cell(row=row, column=5).alignment = align_left
        # ヒントを薄い色で
        if hint != "自由記述":
            ws.cell(row=row, column=5, value=hint).font = Font(name="Yu Gothic", size=9, color="BBBBBB")
        row += 1

    row += 1

    # ============================================================
    # セクション6: 利用環境
    # ============================================================
    row = _write_section_header(ws, row, "6. 利用環境", required=False, badge="確認")

    env_items = [
        ("事務所のインターネット環境", "Wi-Fi / 有線LAN / モバイルルーター / なし"),
        ("運転者が使用する端末", "スマートフォン / タブレット / 既存端末あり / 新規購入予定"),
        ("端末のOS", "iPhone / Android / 両方"),
        ("ブラウザ", "Chrome / Safari / その他"),
    ]
    for item, hint in env_items:
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=item).font = font_normal
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4).border = thin_border
        ws.merge_cells(f"E{row}:G{row}")
        ws.cell(row=row, column=5, value=hint).font = Font(name="Yu Gothic", size=9, color="BBBBBB")
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=7).border = thin_border
        row += 1

    row += 2

    # === フッター ===
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value="ご記入ありがとうございます。内容を確認の上、環境セットアップを進めさせていただきます。").font = font_note
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value="株式会社Smiley　担当: 杉原").font = Font(name="Yu Gothic", size=10, bold=True, color="333333")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    # === 印刷設定 ===
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.8
    ws.page_margins.bottom = 0.6

    # 保存
    output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs", "hearing-sheet.xlsx")
    # ファイルがロックされている場合は別名で保存
    try:
        wb.save(output_path)
    except PermissionError:
        output_path = output_path.replace(".xlsx", "_new.xlsx")
        wb.save(output_path)
    print(f"生成完了: {output_path}")
    return output_path


def _write_section_header(ws, row, title, required=True, badge=None):
    """セクションヘッダーを書き込む"""
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28

    # バッジ
    ws.merge_cells(f"F{row}:G{row}")
    badge_cell = ws.cell(row=row, column=6)
    if badge == "確認":
        badge_cell.value = "確認"
        badge_cell.font = font_required
        badge_cell.fill = fill_blue
    elif required:
        badge_cell.value = "必須"
        badge_cell.font = font_required
        badge_cell.fill = fill_required
    else:
        badge_cell.value = "任意"
        badge_cell.font = font_optional
        badge_cell.fill = fill_optional
    badge_cell.alignment = align_center
    badge_cell.border = thin_border

    return row + 1


def _write_form_row(ws, row, label, placeholder=""):
    """ラベル＋入力欄の行を書き込む"""
    ws.merge_cells(f"B{row}:C{row}")
    cell = ws.cell(row=row, column=2, value=label)
    cell.font = font_header
    cell.fill = fill_header
    cell.border = thin_border
    cell.alignment = align_left
    ws.cell(row=row, column=3).border = thin_border

    ws.merge_cells(f"D{row}:G{row}")
    input_cell = ws.cell(row=row, column=4)
    if placeholder:
        input_cell.value = placeholder
        input_cell.font = Font(name="Yu Gothic", size=9, color="BBBBBB")
    else:
        input_cell.font = font_input
    input_cell.border = thin_border
    input_cell.alignment = align_left
    for c in [5, 6, 7]:
        ws.cell(row=row, column=c).border = thin_border

    return row + 1


def _write_note(ws, row, text):
    """注釈テキストを書き込む"""
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = font_note
    cell.alignment = align_left
    return row + 1


if __name__ == "__main__":
    create_hearing_sheet()
